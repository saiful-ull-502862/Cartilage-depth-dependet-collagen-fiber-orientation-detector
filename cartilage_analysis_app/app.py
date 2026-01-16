import os
import io
import base64
import numpy as np
import cv2
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook

# Attempt to import the AI model
try:
    from ai_model import ZoneDetector
    # Initialize global detector (loads ResNet50)
    print("Initializing AI Model...")
    detector = ZoneDetector()
    AI_AVAILABLE = True
    print("AI Model Initialized Successfully.")
except Exception as e:
    print(f"AI Model not available: {e}")
    AI_AVAILABLE = False
    detector = None

app = Flask(__name__)

def hue_to_angle(hue, zero_hue=0, ninety_hue=60):
    """
    Convert OpenCV HSV Hue (0-179) to fiber orientation angle (0-90 degrees)
    using dynamic anchors.
    
    Args:
        hue: input hue value or array
        zero_hue: Hue value corresponding to 0 degrees (SZ)
        ninety_hue: Hue value corresponding to 90 degrees (DZ)
    """
    # Avoid division by zero
    if abs(ninety_hue - zero_hue) < 1.0:
        # Fallback to defaults or return 0
        denom = 60.0
    else:
        denom = ninety_hue - zero_hue
        
    angle = (hue - zero_hue) / denom * 90.0
    return np.clip(angle, 0, 90)

def analyze_zone(image_zone, zero_hue=0, ninety_hue=60):
    """
    Analyze a specific zone of the image using dynamic HSV hue-to-angle mapping.
    """
    if image_zone.size == 0:
         return {
            "avg_color_hex": "#000000",
            "angle_histogram": [0]*91,
            "angle_labels": list(range(91)),
            "mean_angle": 0,
            "std_angle": 0,
            "avg_hue": 0,
            "avg_r": 0,
            "avg_g": 0,
            "avg_b": 0,
            "avg_intensity": 0
        }

    # Convert to HSV
    hsv = cv2.cvtColor(image_zone, cv2.COLOR_BGR2HSV)
    h_channel = hsv[:, :, 0].flatten()
    s_channel = hsv[:, :, 1].flatten()
    v_channel = hsv[:, :, 2].flatten()
    
    # Masking
    mask = (v_channel > 10) & (s_channel > 10) # Lowered threshold to include dark MZ
    
    valid_hues = h_channel[mask]
    valid_values = v_channel[mask]
    
    if len(valid_hues) == 0:
         return {
            "avg_color_hex": "#000000",
            "angle_histogram": [0]*91,
            "angle_labels": list(range(91)),
            "mean_angle": 0,
            "std_angle": 0,
            "avg_hue": 0,
            "avg_r": 0,
            "avg_g": 0,
            "avg_b": 0,
            "avg_intensity": 0
        }
    
    # Convert hues to angles using DYNAMIC mapping
    angles = hue_to_angle(valid_hues.astype(float), zero_hue, ninety_hue)
    
    # Calculate Histogram
    hist_angle, _ = np.histogram(angles, bins=91, range=(0, 91))
    
    # Normalize
    hist_sum = hist_angle.sum()
    if hist_sum > 0:
        hist_angle = hist_angle / hist_sum
        
    # Calculate Average Color
    pixels = image_zone.reshape(-1, 3)
    valid_pixels = pixels[mask]
    
    if len(valid_pixels) > 0:
        avg_b = np.mean(valid_pixels[:, 0])
        avg_g = np.mean(valid_pixels[:, 1])
        avg_r = np.mean(valid_pixels[:, 2])
        avg_hex = '#{:02x}{:02x}{:02x}'.format(int(avg_r), int(avg_g), int(avg_b))
    else:
        avg_r, avg_g, avg_b = 0, 0, 0
        avg_hex = '#000000'
    
    avg_hue = float(np.mean(valid_hues))
    avg_intensity = float(np.mean(valid_values))
    
    return {
        "avg_color_hex": avg_hex,
        "angle_histogram": hist_angle.tolist(),
        "angle_labels": list(range(91)),
        "mean_angle": float(np.mean(angles)),
        "std_angle": float(np.std(angles)),
        "avg_hue": avg_hue,
        "avg_r": int(avg_r),
        "avg_g": int(avg_g),
        "avg_b": int(avg_b),
        "avg_intensity": avg_intensity
    }

@app.route('/')
def index():
    return render_template('index.html')



@app.route('/reference_batch')
def reference_batch():
    return render_template('reference_batch.html')

@app.route('/scientific_view')
def scientific_view():
    return render_template('scientific_view.html')

@app.route('/analyze_scientific', methods=['POST'])
def analyze_scientific():
    try:
        data = request.json
        if not data or 'image' not in data:
            return jsonify({'error': 'No image data provided'}), 400
        
        # Decode base64 
        image_data = data['image'].split(',')[1]
        sz_b_pct = int(data.get('sz_boundary', 10))
        mz_b_pct = int(data.get('mz_boundary', 30))
        image_bytes = base64.b64decode(image_data)
        nparr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            return jsonify({'error': 'Failed to decode image'}), 400

        # 1. Processing
        hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
        h_chan = hsv[:, :, 0]
        s_chan = hsv[:, :, 1]
        v_chan = hsv[:, :, 2]
        
        # 2. Masking (V > 10 and S > 10)
        mask = (v_chan > 10) & (s_chan > 10)
        mask_uint8 = (mask * 255).astype(np.uint8)
        
        # 3. Create Visualizations
        
        # A. Hue Channel (Heatmap)
        # Apply colormap to Hue channel for better visualization
        # Normalize H (0-179) to 0-255
        h_norm = cv2.normalize(h_chan, None, 0, 255, cv2.NORM_MINMAX)
        hue_heatmap = cv2.applyColorMap(h_norm, cv2.COLORMAP_JET)
        
        # B. Mask Visualization
        # Show original image where mask is valid, Black otherwise
        masked_img = cv2.bitwise_and(img, img, mask=mask_uint8)
        
        # C. Zone Overlay Visualization
        zone_img = img.copy()
        h, w = img.shape[:2]
        # Calculate Y coords
        y_sz = int(h * (sz_b_pct / 100.0))
        y_mz = int(h * (mz_b_pct / 100.0))
        
        # Draw Lines
        cv2.line(zone_img, (0, y_sz), (w, y_sz), (0, 255, 255), 2) # Yellow
        cv2.line(zone_img, (0, y_mz), (w, y_mz), (0, 255, 255), 2)
        cv2.putText(zone_img, "SZ", (10, min(y_sz - 10, 30)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,255), 2)
        cv2.putText(zone_img, "MZ", (10, min(y_mz - 10, y_sz+30)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,255), 2)
        cv2.putText(zone_img, "DZ", (10, min(h - 10, y_mz+30)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,255), 2)

        # Save images to static/uploads
        def save_temp_img(image_array, prefix):
            filename = f"{prefix}_{np.random.randint(10000, 99999)}.jpg"
            filepath = os.path.join('static', 'uploads', filename)
            cv2.imwrite(filepath, image_array)
            return f"/static/uploads/{filename}"
            
        original_url = save_temp_img(img, "sc_orig")
        hue_url = save_temp_img(hue_heatmap, "sc_hue")
        mask_url = save_temp_img(masked_img, "sc_mask")
        zone_url = save_temp_img(zone_img, "sc_zones")
        
        # 4. Histogram Calculation (ONLY Valid Pixels)
        valid_hues = h_chan[mask]
        # Histogram 0-179
        hist_counts, _ = np.histogram(valid_hues, bins=180, range=(0, 180))
        
        return jsonify({
            'success': True,
            'original_url': original_url,
            'hue_url': hue_url,
            'mask_url': mask_url,
            'zone_url': zone_url,
            'histogram': hist_counts.tolist()
        })

    except Exception as e:
        print(f"Scientific Analysis Error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/analyze', methods=['POST'])
def analyze():
    try:
        data = request.json
        if not data or 'image' not in data:
            return jsonify({'error': 'No image data provided'}), 400
        
        # Decode base64 
        image_data = data['image'].split(',')[1] 
        image_bytes = base64.b64decode(image_data)
        nparr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            return jsonify({'error': 'Failed to decode image'}), 400

        height, width, _ = img.shape
        
        # Default Mapping
        zero_hue = 0   # Red
        ninety_hue = 60 # Green

        # Check for overrides
        force_zero = data.get('force_zero_hue')
        force_ninety = data.get('force_ninety_hue')
        
        if force_zero is not None and force_ninety is not None:
             zero_hue = float(force_zero)
             ninety_hue = float(force_ninety)
             print(f"Using forced color calibration: 0°={zero_hue}, 90°={ninety_hue}")

        
        # AI Auto-Detection Logic
        # New flag from frontend
        use_ai = data.get('use_ai', False)
        ai_result_info = {}
        
        split1_pct = float(data.get('sz_boundary', 33))
        split2_pct = float(data.get('mz_boundary', 66))

        if use_ai and AI_AVAILABLE and detector:
            print("Running AI Detection...")
            ai_results = detector.detect_zones_and_colors(img)
            
            if ai_results.get('success'):
                # Update boundaries
                split1_pct = ai_results['sz_boundary']
                split2_pct = ai_results['mz_boundary']
                
                # Update Color Mapping
                detected_sz_hue = ai_results['sz_hue']
                detected_dz_hue = ai_results['dz_hue']
                
                print(f"AI Detected: SZ Bound={split1_pct}%, MZ Bound={split2_pct}%")
                print(f"AI Detected Colors: SZ Hue={detected_sz_hue:.1f}, DZ Hue={detected_dz_hue:.1f}")
                
                # Update logic if separation is sufficient
                # Update logic if separation is sufficient AND we aren't using forced values
                if abs(detected_dz_hue - detected_sz_hue) > 5:
                    # Only update if NOT forced
                    if force_zero is None or force_ninety is None:
                        zero_hue = detected_sz_hue
                        ninety_hue = detected_dz_hue
                    else:
                        print("AI detected colors ignored in favor of forced calibration.")
                
                ai_result_info = {
                    'detected': True,
                    'sz_hue': round(detected_sz_hue, 1),
                    'dz_hue': round(detected_dz_hue, 1),
                    'sz_boundary': split1_pct,
                    'mz_boundary': split2_pct
                }
            else:
                print("AI Detection returned failure.")
        
        # Ensure splits are sorted and within bounds
        s1 = np.clip(split1_pct / 100.0, 0.0, 1.0)
        s2 = np.clip(split2_pct / 100.0, 0.0, 1.0)
        if s1 > s2: s1 = s2
        
        if not use_ai and force_zero is None and force_ninety is None:
             print(f"Manual Recalculation: Detecting colors from zones s1={s1}, s2={s2}")
             # Get SZ ROI
             h_px, w_px, _ = img.shape
             
             # Helper for circular mean
             def circular_hue_mean(h_values):
                 # Hues are 0-179 in OpenCV. Convert to radians 0-2pi
                 rads = np.deg2rad(h_values * 2.0)
                 sin_sum = np.sum(np.sin(rads))
                 cos_sum = np.sum(np.cos(rads))
                 mean_rad = np.arctan2(sin_sum, cos_sum)
                 mean_deg = np.rad2deg(mean_rad)
                 if mean_deg < 0: mean_deg += 360
                 return mean_deg / 2.0

             # Calculate SZ Hue (0 to s1)
             z1_h = int(h_px * s1)
             if z1_h > 0:
                 roi_sz = img[0:z1_h, :]
                 hsv_sz = cv2.cvtColor(roi_sz, cv2.COLOR_BGR2HSV)
                 h_sz = hsv_sz[:, :, 0]
                 v_sz = hsv_sz[:, :, 2]
                 mask_sz = v_sz > 20
                 if np.any(mask_sz):
                     zero_hue = circular_hue_mean(h_sz[mask_sz])
            
             # Calculate DZ Hue (s2 to 1.0)
             z2_h = int(h_px * s2)
             if z2_h < h_px:
                 roi_dz = img[z2_h:h_px, :]
                 hsv_dz = cv2.cvtColor(roi_dz, cv2.COLOR_BGR2HSV)
                 h_dz = hsv_dz[:, :, 0]
                 v_dz = hsv_dz[:, :, 2]
                 mask_dz = v_dz > 20
                 if np.any(mask_dz):
                     ninety_hue = circular_hue_mean(h_dz[mask_dz])
             
             print(f"Manually Detected: Zero={zero_hue:.1f}, Ninety={ninety_hue:.1f}") 
        
        z1_h = int(height * s1)
        z2_h = int(height * s2)
        
        zone_sz = img[0:z1_h, :]
        zone_mz = img[z1_h:z2_h, :]
        zone_dz = img[z2_h:height, :]
        
        results = {
            "SZ": analyze_zone(zone_sz, zero_hue, ninety_hue),
            "MZ": analyze_zone(zone_mz, zero_hue, ninety_hue),
            "DZ": analyze_zone(zone_dz, zero_hue, ninety_hue)
        }

        # Create Annotated Image
        annotated_img = img.copy()
        line_color = (0, 255, 255) 
        thickness = 2
        cv2.line(annotated_img, (0, z1_h), (width, z1_h), line_color, thickness)
        cv2.line(annotated_img, (0, z2_h), (width, z2_h), line_color, thickness)
        
        font = cv2.FONT_HERSHEY_SIMPLEX
        cv2.putText(annotated_img, "SZ", (10, min(z1_h - 10, 30)), font, 0.8, (255,255,255), 2)
        cv2.putText(annotated_img, "MZ", (10, min(z2_h - 10, z1_h + 30)), font, 0.8, (255,255,255), 2)
        cv2.putText(annotated_img, "DZ", (10, min(height - 10, z2_h + 30)), font, 0.8, (255,255,255), 2)
        
        filename = f"analyzed_{np.random.randint(1000, 9999)}.jpg"
        filepath = os.path.join('static', 'uploads', filename)
        cv2.imwrite(filepath, annotated_img)
        
        # Calculate Depth Profile
        bins = 100
        raw_profile = []
        
        for i in range(bins):
            t_start = i / float(bins)
            t_end = (i + 1) / float(bins)
            
            # Top is 0, Bottom is 1
            y_start = int(height * t_start)
            y_end = int(height * t_end)
            
            y_start = max(0, y_start)
            y_end = min(height, y_end)
            
            if y_end <= y_start:
                 # Fallback for very small images or rounding
                 y_end = y_start + 1
            
            roi = img[y_start:y_end, :]
            
            # Analyze this slice
            hsv_roi = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
            h_roi = hsv_roi[:, :, 0]
            v_roi = hsv_roi[:, :, 2]
            
            mask = v_roi > 20
            
            if np.any(mask):
                 valid_hues = h_roi[mask]
                 
                 # Circular mean for hue
                 rads = np.deg2rad(valid_hues * 2.0)
                 sin_sum = np.sum(np.sin(rads))
                 cos_sum = np.sum(np.cos(rads))
                 mean_rad = np.arctan2(sin_sum, cos_sum)
                 mean_hue = np.rad2deg(mean_rad)
                 if mean_hue < 0: mean_hue += 360
                 mean_hue /= 2.0
                 
                 # Colors
                 avg_b = np.mean(roi[:,:,0][mask])
                 avg_g = np.mean(roi[:,:,1][mask])
                 avg_r = np.mean(roi[:,:,2][mask])
                 intensity = np.mean(v_roi[mask])
                 
                 # Std Dev (Angle)
                 # Map hue to angle first, then std? 
                 # Angle = (Hue - Zero) / (Ninety - Zero) * 90
                 # For simplicity, calculate std of hues then convert scale
                 std_hue = np.std(valid_hues)
                 scale_factor = 90.0 / (abs(ninety_hue - zero_hue) if abs(ninety_hue - zero_hue) > 1 else 1.0)
                 std_angle = std_hue * scale_factor
                 
                 # Angle Calculation
                 # Helper to map hue to angle
                 def hue_to_angle(h, z, n):
                      # Handle wrapping if needed, but simple linear for now as per previous logic
                      # dist logic:
                      # if z < n: 0 at z, 90 at n
                      total_range = n - z
                      if abs(total_range) < 0.1: return 0
                      
                      curr = h - z
                      ratio = curr / total_range
                      ang = ratio * 90.0
                      return np.clip(ang, 0, 90)

                 angle = hue_to_angle(mean_hue, zero_hue, ninety_hue)
                 
                 # Hex
                 def rgb_to_hex(r, g, b):
                     return "#{:02x}{:02x}{:02x}".format(int(r), int(g), int(b))
                     
                 raw_profile.append({
                     'thickness': round((t_start + t_end)/2, 3),
                     'angle': round(angle, 2),
                     'std': round(std_angle, 2),
                     'mean_hue': round(mean_hue, 1),
                     'avg_r': int(avg_r),
                     'avg_g': int(avg_g),
                     'avg_b': int(avg_b),
                     'avg_hex': rgb_to_hex(avg_r, avg_g, avg_b),
                     'intensity': int(intensity)
                 })
            else:
                 # Empty bin
                 raw_profile.append(None)

        # Interpolation to fill gaps
        filled_profile = []
        for i in range(bins):
            item = raw_profile[i]
            thickness = round(((i / float(bins)) + ((i + 1) / float(bins))) / 2, 3)
            
            if item is None:
                # Find previous valid
                prev_valid = None
                for k in range(i-1, -1, -1):
                    if raw_profile[k] is not None:
                        prev_valid = raw_profile[k]
                        break
                
                # Find next valid
                next_valid = None
                for k in range(i+1, bins):
                    if raw_profile[k] is not None:
                        next_valid = raw_profile[k]
                        break
                
                # Interpolate
                if prev_valid and next_valid:
                     # Simple linear average for middle gap
                     # Improve to proper weighted interp if needed, but average is safe for small gaps
                     fill_item = {
                         'thickness': thickness,
                         'angle': round((prev_valid['angle'] + next_valid['angle'])/2, 2),
                         'std': round((prev_valid['std'] + next_valid['std'])/2, 2),
                         'mean_hue': round((prev_valid['mean_hue'] + next_valid['mean_hue'])/2, 1),
                         'avg_r': int((prev_valid['avg_r'] + next_valid['avg_r'])/2),
                         'avg_g': int((prev_valid['avg_g'] + next_valid['avg_g'])/2),
                         'avg_b': int((prev_valid['avg_b'] + next_valid['avg_b'])/2),
                         'avg_hex': prev_valid['avg_hex'], # Just reuse hex
                         'intensity': int((prev_valid['intensity'] + next_valid['intensity'])/2)
                     }
                     filled_profile.append(fill_item)
                elif prev_valid:
                    # propagate forward
                    new_item = prev_valid.copy()
                    new_item['thickness'] = thickness
                    filled_profile.append(new_item)
                elif next_valid:
                    # propagate backward
                    new_item = next_valid.copy()
                    new_item['thickness'] = thickness
                    filled_profile.append(new_item)
                else:
                     # No data at all?
                     filled_profile.append({
                         'thickness': thickness,
                         'angle': 0, 'std': 0, 'mean_hue': 0, 
                         'avg_r':0, 'avg_g':0, 'avg_b':0, 'avg_hex':'#000000', 'intensity':0
                     })
            else:
                filled_profile.append(item)

        depth_profile = filled_profile
            
        zone_boundaries = {
            'sz_end': round(s1, 3), # Top boundary (0 to s1)
            'mz_end': round(s2, 3), # Middle boundary (s1 to s2)
            'sz_boundary': round(s1, 3), 
            'mz_boundary': round(s2, 3)
        }
            
        return jsonify({
            'success': True, 
            'results': results, 
            'annotated_image_url': f'/static/uploads/{filename}',
            'depth_profile': depth_profile,
            'zone_boundaries': zone_boundaries,
            'ai_info': ai_result_info,
            'color_calibration': {
                'zero_hue': zero_hue,
                'ninety_hue': ninety_hue
            }
        })
        
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/download_excel', methods=['POST'])
def download_excel():
    """Generate and download Excel file with depth profile data."""
    try:
        data = request.json
        depth_profile = data.get('depth_profile', [])
        zone_boundaries = data.get('zone_boundaries', {})
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Depth Profile"
        
        # Headers
        ws['A1'] = 'Normalized Thickness'
        ws['B1'] = 'Mean Angle (Degrees)'
        ws['C1'] = 'Std Dev'
        ws['D1'] = 'Zone'
        ws['E1'] = 'Avg Hue (OpenCV)'
        ws['F1'] = 'Red (R)'
        ws['G1'] = 'Green (G)'
        ws['H1'] = 'Blue (B)'
        ws['I1'] = 'Hex Color'
        ws['J1'] = 'Intensity (V)'
        
        # Zone boundaries - logic: 0 is top
        # SZ is 0 to sz_boundary
        sz_end = zone_boundaries.get('sz_boundary', 0.33)
        mz_end = zone_boundaries.get('mz_boundary', 0.66)
        
        # Data rows
        for i, item in enumerate(depth_profile, start=2):
            ws[f'A{i}'] = item.get('thickness')
            ws[f'B{i}'] = item.get('angle')
            ws[f'C{i}'] = item.get('std')
            
            # Determine zone
            t = item.get('thickness', 0)
            if t <= sz_end:
                zone = 'SZ'
            elif t <= mz_end:
                zone = 'MZ'
            else:
                zone = 'DZ'
            ws[f'D{i}'] = zone
            
            # Color information
            ws[f'E{i}'] = item.get('avg_hue')
            ws[f'F{i}'] = item.get('avg_r')
            ws[f'G{i}'] = item.get('avg_g')
            ws[f'H{i}'] = item.get('avg_b')
            ws[f'I{i}'] = item.get('avg_hex')
            ws[f'J{i}'] = item.get('intensity')
        
        # Add explanation sheet
        ws2 = wb.create_sheet("Color Mapping Info")
        ws2['A1'] = "Color to Angle Mapping Explanation"
        ws2['A3'] = "This analysis uses the HSV (Hue, Saturation, Value) color space."
        ws2['A4'] = "The Hue channel (0-179 in OpenCV) is mapped to fiber orientation angles (0-90 degrees)."
        
        ws2['A6'] = "Mapping Logic (Standard/Dynamic):"
        ws2['A7'] = "Linear interpolation between SZ Color (0 deg) and DZ Color (90 deg)."
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='fibril_angle_depth_profile.xlsx'
        )
        
    except Exception as e:
        print(f"Excel Error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/batch')
def batch():
    """Batch analysis page for multiple images."""
    return render_template('batch.html')

@app.route('/batch_download_excel', methods=['POST'])
def batch_download_excel():
    """Generate and download Excel file with batch analysis results."""
    try:
        data = request.json
        results = data.get('results', [])
        ref_data = data.get('reference_data')
        
        if not results and not ref_data:
            return jsonify({'error': 'No results provided'}), 400
        
        wb = Workbook()
        
        # --- REFERENCE SHEET (First if exists) ---
        if ref_data:
            ws_ref = wb.active
            ws_ref.title = "Reference Profile"
            
            # Metadata
            ws_ref['A1'] = "Reference Analysis Data"
            ws_ref['A2'] = f"0-Degree Hue: {ref_data.get('color_calibration',{}).get('zero_hue','N/A')}"
            ws_ref['A3'] = f"90-Degree Hue: {ref_data.get('color_calibration',{}).get('ninety_hue','N/A')}"
            ws_ref['B1'] = f"SZ Boundary: {ref_data.get('zone_boundaries',{}).get('sz_boundary',0)*100}%"
            ws_ref['B2'] = f"MZ Boundary: {ref_data.get('zone_boundaries',{}).get('mz_boundary',0)*100}%"
            
            # Headers
            headers = ['Normalized Thickness', 'Angle', 'Std Dev', 'Zone', 'Avg Hue', 'Intensity']
            for c, h in enumerate(headers, 1):
                ws_ref.cell(row=5, column=c, value=h)
            
            # Ref Boundaries for Zone calc
            rb_sz = ref_data.get('zone_boundaries',{}).get('sz_boundary', 0.33)
            rb_mz = ref_data.get('zone_boundaries',{}).get('mz_boundary', 0.66)
            
            # Data
            for idx, dp in enumerate(ref_data.get('depth_profile', []), start=6):
                t = dp.get('thickness', 0)
                ws_ref.cell(row=idx, column=1, value=t)
                ws_ref.cell(row=idx, column=2, value=dp.get('angle'))
                ws_ref.cell(row=idx, column=3, value=dp.get('std'))

                
                # Zone Logic
                z = 'Unknown'
                if t <= rb_sz: z='SZ'
                elif t <= rb_mz: z='MZ'
                else: z='DZ'
                
                ws_ref.cell(row=idx, column=4, value=z)
                ws_ref.cell(row=idx, column=5, value=dp.get('mean_hue'))
                ws_ref.cell(row=idx, column=6, value=dp.get('mean_intensity'))
                
            # Create next sheet for Summary
            ws1 = wb.create_sheet("Batch Summary")
        else:
            ws1 = wb.active
            ws1.title = "Summary"
        
        # --- BATCH SUMMARY ---
        ws1['A1'] = 'Filename'
        ws1['B1'] = 'SZ Mean Angle'
        ws1['C1'] = 'SZ Std Dev'
        ws1['D1'] = 'MZ Mean Angle'
        ws1['E1'] = 'MZ Std Dev'
        ws1['F1'] = 'DZ Mean Angle'
        ws1['G1'] = 'DZ Std Dev'
        
        sz_means = []
        mz_means = []
        dz_means = []
        
        for i, result in enumerate(results, start=2):
            ws1[f'A{i}'] = result.get('filename', f'Image {i-1}')
            
            sz_data = result.get('results', {}).get('SZ', {})
            mz_data = result.get('results', {}).get('MZ', {})
            dz_data = result.get('results', {}).get('DZ', {})
            
            sz_mean = sz_data.get('mean_angle', 0)
            mz_mean = mz_data.get('mean_angle', 0)
            dz_mean = dz_data.get('mean_angle', 0)
            
            ws1[f'B{i}'] = round(sz_mean, 2)
            ws1[f'C{i}'] = round(sz_data.get('std_angle', 0), 2)
            ws1[f'D{i}'] = round(mz_mean, 2)
            ws1[f'E{i}'] = round(mz_data.get('std_angle', 0), 2)
            ws1[f'F{i}'] = round(dz_mean, 2)
            ws1[f'G{i}'] = round(dz_data.get('std_angle', 0), 2)
            
            sz_means.append(sz_mean)
            mz_means.append(mz_mean)
            dz_means.append(dz_mean)
        
        # Add summary rows
        row = len(results) + 2
        ws1[f'A{row}'] = 'MEAN'
        ws1[f'B{row}'] = round(np.mean(sz_means), 2) if sz_means else 0
        ws1[f'D{row}'] = round(np.mean(mz_means), 2) if mz_means else 0
        ws1[f'F{row}'] = round(np.mean(dz_means), 2) if dz_means else 0
        
        row += 1
        ws1[f'A{row}'] = 'STD DEV'
        ws1[f'B{row}'] = round(np.std(sz_means), 2) if sz_means else 0
        ws1[f'D{row}'] = round(np.std(mz_means), 2) if mz_means else 0
        ws1[f'F{row}'] = round(np.std(dz_means), 2) if dz_means else 0
        
        # Sheet 2: Combined Depth Profile
        ws2 = wb.create_sheet("Combined Depth Profile")
        ws2['A1'] = 'Normalized Thickness'
        
        from openpyxl.utils import get_column_letter

        # Add columns for each image
        for i, result in enumerate(results):
            col = get_column_letter(i + 2)
            ws2[f'{col}1'] = result.get('filename', f'Image {i+1}')
        
        # Add Mean and Std columns
        mean_col_idx = len(results) + 2
        std_col_idx = len(results) + 3
        ws2.cell(row=1, column=mean_col_idx, value='Mean')
        ws2.cell(row=1, column=std_col_idx, value='Std Dev')
        
        # Normalize all depth profiles to 100 bins
        bins = 100
        all_profiles = []
        
        for result in results:
            profile = [None] * bins
            for dp in result.get('depth_profile', []):
                idx = int(round(dp.get('thickness', 0) * (bins - 1)))
                if 0 <= idx < bins:
                    profile[idx] = dp.get('angle')
            all_profiles.append(profile)
        
        # Fill in data
        for bin_idx in range(bins):
            row = bin_idx + 2
            thickness = bin_idx / (bins - 1)
            ws2.cell(row=row, column=1, value=round(thickness, 3))
            
            values = []
            for i, profile in enumerate(all_profiles):
                val = profile[bin_idx]
                ws2.cell(row=row, column=i+2, value=round(val, 2) if val is not None else None)
                if val is not None:
                    values.append(val)
            
            # Calculate mean and std
            if values:
                ws2.cell(row=row, column=mean_col_idx, value=round(np.mean(values), 2))
                ws2.cell(row=row, column=std_col_idx, value=round(np.std(values), 2))
        
        # Save to BytesIO
        output = io.BytesIO()
        
        # Add detailed sheets for EACH image
        for result in results:
            fname = result.get('filename', 'Unknown')
            # Clean filename for Excel sheet naming (max 31 chars, no illegal chars)
            clean_name = "".join([c for c in fname if c.isalnum() or c in (' ','_','-')]) 
            sheet_name = clean_name[:30] if len(clean_name)>30 else clean_name
            
            ws_detail = wb.create_sheet(sheet_name)
            
            # Match Reference Excel Format Headers
            # Format: Thickness, Mean Angle, Std Dev, Zone, Hue, R, G, B, Hex, Value
            headers = ['Normalized Thickness', 'Mean Angle (Degrees)', 'Std Dev', 'Zone', 'Avg Hue', 'Red', 'Green', 'Blue', 'Hex', 'Intensity']
            for c, h in enumerate(headers, 1):
                ws_detail.cell(row=1, column=c, value=h)
            
            sz_b = result.get('zone_boundaries',{}).get('sz_boundary', 0.33)
            mz_b = result.get('zone_boundaries',{}).get('mz_boundary', 0.66)
            
            # Rows
            for idx, dp in enumerate(result.get('depth_profile', []), start=2):
                t = dp.get('thickness', 0)
                ws_detail.cell(row=idx, column=1, value=t)
                ws_detail.cell(row=idx, column=2, value=dp.get('angle'))
                ws_detail.cell(row=idx, column=3, value=dp.get('std'))
                
                if t <= sz_b: z='SZ'
                elif t <= mz_b: z='MZ'
                else: z='DZ'
                     
                ws_detail.cell(row=idx, column=4, value=z)
                ws_detail.cell(row=idx, column=5, value=dp.get('mean_hue')) # Avg Hue
                
                # Colors might not be in depth profile for batch? 
                # analyze_zone -> get_depth_profile currently only packs: thickness, angle, std, mean_hue, mean_intensity
                # We should try to pack RGB if available, but if not put placeholders
                
                ws_detail.cell(row=idx, column=6, value=dp.get('avg_r', 0)) # Red
                ws_detail.cell(row=idx, column=7, value=dp.get('avg_g', 0)) # Green
                ws_detail.cell(row=idx, column=8, value=dp.get('avg_b', 0)) # Blue
                ws_detail.cell(row=idx, column=9, value=dp.get('avg_hex', '#000000')) # Hex
                ws_detail.cell(row=idx, column=10, value=dp.get('intensity', dp.get('mean_intensity', 0))) # Intensity


        wb.save(output)
        output.seek(0)

        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='batch_analysis_results.xlsx'
        )
        
    except Exception as e:
        print(f"Batch Excel Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/convert_image', methods=['POST'])
def convert_image():
    """Convert uploaded image (e.g. TIFF) to PNG base64 for browser display."""
    try:
        file = request.files.get('image')
        if not file:
            return jsonify({'error': 'No file uploaded'}), 400
            
        # Read image to memory
        in_memory_file = io.BytesIO()
        file.save(in_memory_file)
        in_memory_file.seek(0)
        file_bytes = np.asarray(bytearray(in_memory_file.read()), dtype=np.uint8)
        
        # Decode using OpenCV
        img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
        if img is None:
             return jsonify({'error': 'Failed to decode image'}), 400
             
        # Encode as PNG
        _, buffer = cv2.imencode('.png', img)
        png_base64 = base64.b64encode(buffer).decode('utf-8')
        
        return jsonify({
            'success': True,
            'image_url': f"data:image/png;base64,{png_base64}"
        })
        
    except Exception as e:
        print(f"Conversion Error: {e}")
        return jsonify({'error': str(e)}), 500
@app.route('/download_excel_origin', methods=['POST'])
def download_excel_origin():
    """Generate Excel file formatted specifically for OriginPro (X Y XErr)."""
    try:
        data = request.json
        depth_profile = data.get('depth_profile', [])
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Origin Data"
        
        # Origin Format: X(Angle), Y(Thickness), XErr(StdDev)
        # Note: We swap col position. Angle is X, Thickness is Y.
        ws['A1'] = 'Mean Angle (X)'
        ws['B1'] = 'Norm Thickness (Y)'
        ws['C1'] = 'Std Dev (XErr)'
        ws['D1'] = 'Zone' # Extra info
        
        sz_end = data.get('zone_boundaries', {}).get('sz_end', 0.9)
        mz_end = data.get('zone_boundaries', {}).get('mz_end', 0.7)
        
        for i, item in enumerate(depth_profile, start=2):
            angle = item.get('angle')
            width = item.get('thickness')
            std = item.get('std')
            
            if angle is None: continue
            
            ws[f'A{i}'] = angle
            ws[f'B{i}'] = width
            ws[f'C{i}'] = std
            
            # Helper for zone coloring in Origin if needed
            t = width
            if t >= sz_end: zone = 'SZ'
            elif t >= mz_end: zone = 'MZ'
            else: zone = 'DZ'
            ws[f'D{i}'] = zone
            
        # Add comment about column designation
        ws['F1'] = "Origin Import Instructions:"
        ws['F2'] = "1. Import this Sheet"
        ws['F3'] = "2. Select Col A, B, C"
        ws['F4'] = "3. Right Click > Set As > XYYErr"
        ws['F5'] = "4. Plot > Scatter"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='origin_ready_data.xlsx'
        )

    except Exception as e:
        print(f"Origin Excel Error: {e}")
        return jsonify({'error': str(e)}), 500



if __name__ == '__main__':
    app.run(debug=True, port=5001, host='0.0.0.0')

